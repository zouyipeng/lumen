import json
import shutil
from pathlib import Path

from langchain_core.tools import tool

from config import resolve_project_path


def _col_to_index(col: str | int) -> int:
    """Convert Excel column letter (A) or 1-based index to 1-based column index."""
    if isinstance(col, int):
        return col
    if isinstance(col, str) and col.isdigit():
        return int(col)
    col = str(col).upper()
    index = 0
    for ch in col:
        if not ("A" <= ch <= "Z"):
            raise ValueError(f"Invalid column: {col}")
        index = index * 26 + (ord(ch) - ord("A") + 1)
    return index


def _cell_value(value) -> str:
    if value is None:
        return ""
    return str(value)


def _default_layout(rn_column_ids: list[str] | None = None) -> dict:
    """Build default layout when template_path is not configured."""
    rn_column_ids = rn_column_ids or []
    rn_columns = {}
    for idx, col_id in enumerate(rn_column_ids, start=5):
        rn_columns[col_id] = idx
    return {
        "header_row": 1,
        "data_start_row": 2,
        "commit_fields": {
            "short_hash": "A",
            "message": "B",
            "author": "C",
            "date": "D",
        },
        "rn_columns": rn_columns,
    }


def get_excel_layout(excel_config: dict, rn_column_ids: list[str] | None = None) -> dict:
    """Return layout from config or sensible defaults."""
    layout = excel_config.get("layout")
    if layout:
        return layout
    return _default_layout(rn_column_ids)


def copy_template(template_path: str, output_path: str) -> str:
    """Copy template Excel to output path, preserving formatting."""
    src = resolve_project_path(template_path)
    dst = resolve_project_path(output_path)
    if not src.exists():
        return f"错误: 模板文件不存在 - {src}"
    dst.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(src, dst)
    return f"成功复制模板: {src} -> {dst}"


def read_excel_with_layout(
    input_path: str,
    sheet_name: str,
    layout: dict,
    rn_column_defs: list[dict] | None = None,
) -> dict:
    """Read Excel using layout mapping. Returns commits, column_results, last_commit_hash."""
    from openpyxl import load_workbook

    path = resolve_project_path(input_path)
    empty = {"commits": [], "column_results": [], "last_commit_hash": ""}
    if not path.exists():
        return empty

    commit_fields = layout.get("commit_fields", {})
    rn_columns_map = layout.get("rn_columns", {})
    data_start_row = layout.get("data_start_row", 2)

    field_cols = {field: _col_to_index(col) for field, col in commit_fields.items()}
    rn_cols = {col_id: _col_to_index(col) for col_id, col in rn_columns_map.items()}

    wb = load_workbook(str(path), read_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return empty

    ws = wb[sheet_name]
    commits = []
    for row_idx in range(data_start_row, ws.max_row + 1):
        short_hash = _cell_value(ws.cell(row=row_idx, column=field_cols.get("short_hash", 1)).value)
        if not short_hash:
            continue
        commit = {
            "hash": short_hash,
            "short_hash": short_hash,
            "message": _cell_value(ws.cell(row=row_idx, column=field_cols.get("message", 2)).value),
            "author": _cell_value(ws.cell(row=row_idx, column=field_cols.get("author", 3)).value),
            "date": _cell_value(ws.cell(row=row_idx, column=field_cols.get("date", 4)).value),
        }
        commits.append(commit)

    rn_column_defs = rn_column_defs or []
    name_by_id = {c["id"]: c.get("name", c["id"]) for c in rn_column_defs}

    column_results = []
    for col_id, col_idx in rn_cols.items():
        structured = {}
        for row_idx in range(data_start_row, ws.max_row + 1):
            short_hash = _cell_value(ws.cell(row=row_idx, column=field_cols.get("short_hash", 1)).value)
            if not short_hash:
                continue
            value = ws.cell(row=row_idx, column=col_idx).value
            if value is not None and str(value):
                structured[short_hash] = str(value)
        column_results.append({
            "column_id": col_id,
            "column_name": name_by_id.get(col_id, col_id),
            "peon_output": "",
            "structured_result": structured,
            "review_status": "approved",
            "review_feedback": "",
            "retry_count": 0,
        })

    wb.close()
    last_commit_hash = commits[-1]["hash"] if commits else ""
    return {
        "commits": commits,
        "column_results": column_results,
        "last_commit_hash": last_commit_hash,
    }


def write_excel_with_layout(
    output_path: str,
    sheet_name: str,
    layout: dict,
    commits: list[dict],
    column_results: list[dict],
    *,
    metadata: dict | None = None,
    metadata_values: dict | None = None,
    write_metadata: bool = True,
) -> str:
    """Write commits and RN column values into Excel using layout mapping."""
    from openpyxl import Workbook, load_workbook

    path = resolve_project_path(output_path)
    commit_fields = layout.get("commit_fields", {})
    rn_columns_map = layout.get("rn_columns", {})
    data_start_row = layout.get("data_start_row", 2)

    field_cols = {field: _col_to_index(col) for field, col in commit_fields.items()}
    rn_cols = {r["column_id"]: _col_to_index(rn_columns_map[r["column_id"]])
               for r in column_results if r["column_id"] in rn_columns_map}

    results_by_id = {r["column_id"]: r for r in column_results}

    if path.exists():
        wb = load_workbook(str(path))
    else:
        wb = Workbook()

    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.active
        ws.title = sheet_name

    if write_metadata and metadata and metadata_values:
        for key, cell_ref in metadata.items():
            if key in metadata_values and cell_ref:
                ws[cell_ref] = metadata_values[key]

    for offset, commit in enumerate(commits):
        row_idx = data_start_row + offset
        for field, col_idx in field_cols.items():
            ws.cell(row=row_idx, column=col_idx, value=commit.get(field, ""))
        for col_id, col_idx in rn_cols.items():
            result = results_by_id.get(col_id, {})
            structured = result.get("structured_result", {})
            commit_hash = commit.get("hash", "")
            short_hash = commit.get("short_hash", "")
            value = (
                structured.get(commit_hash)
                or structured.get(short_hash)
                or structured.get("all", "")
            )
            ws.cell(row=row_idx, column=col_idx, value=str(value) if value else "")

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))
    wb.close()
    return f"成功写入 Excel: {path} ({len(commits)} 行)"


def write_excel_legacy(data: str, output_path: str, sheet_name: str = "Release Note") -> str:
    """Legacy: write Excel from JSON columns/rows (no template)."""
    try:
        from openpyxl import Workbook

        parsed = json.loads(data)
        columns = parsed.get("columns", [])
        rows = parsed.get("rows", [])

        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        for col_idx, col_name in enumerate(columns, start=1):
            ws.cell(row=1, column=col_idx, value=col_name)

        for row_idx, row_data in enumerate(rows, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=str(value) if value is not None else "")

        for col_idx, col_name in enumerate(columns, start=1):
            max_len = len(str(col_name))
            for row in rows:
                if col_idx - 1 < len(row) and row[col_idx - 1]:
                    max_len = max(max_len, len(str(row[col_idx - 1])))
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 2, 60)

        path = resolve_project_path(output_path)
        path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(path))
        return f"成功写入 Excel: {path} ({len(rows)} 行)"

    except ImportError:
        return "错误: openpyxl 未安装，请运行 pip install openpyxl"
    except json.JSONDecodeError as exc:
        return f"错误: JSON 解析失败 - {exc}"
    except Exception as exc:
        return f"错误: 写入 Excel 失败 - {exc}"


@tool
def write_excel(data: str, output_path: str, sheet_name: str = "Release Note") -> str:
    """将 JSON 数据写入 Excel 文件。data 为 JSON 字符串，格式为 {"columns": ["列1", "列2"], "rows": [["值1", "值2"], ...]}。"""
    return write_excel_legacy(data, output_path, sheet_name)


@tool
def read_excel(input_path: str, sheet_name: str = "Release Note") -> str:
    """读取已有 Excel 文件，返回 JSON 格式数据。格式为 {"columns": ["列1", "列2"], "rows": [["值1", "值2"], ...]}。"""
    try:
        from openpyxl import load_workbook

        path = resolve_project_path(input_path)
        if not path.exists():
            return json.dumps({"columns": [], "rows": []}, ensure_ascii=False)

        wb = load_workbook(str(path), read_only=True)
        if sheet_name not in wb.sheetnames:
            wb.close()
            return json.dumps({"columns": [], "rows": []}, ensure_ascii=False)

        ws = wb[sheet_name]
        rows_iter = ws.iter_rows(values_only=True)

        header_row = next(rows_iter, None)
        if header_row is None:
            wb.close()
            return json.dumps({"columns": [], "rows": []}, ensure_ascii=False)

        columns = [str(c) if c is not None else "" for c in header_row]
        rows = []
        for row in rows_iter:
            rows.append([str(c) if c is not None else "" for c in row])

        wb.close()
        return json.dumps({"columns": columns, "rows": rows}, ensure_ascii=False)

    except ImportError:
        return "错误: openpyxl 未安装，请运行 pip install openpyxl"
    except Exception as exc:
        return f"错误: 读取 Excel 失败 - {exc}"


EXCEL_TOOLS = [write_excel, read_excel]
